"""Utility helpers for composing Excel workbook exports."""

from __future__ import annotations

from io import BytesIO
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CELL_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _sanitize_row(row: Sequence[object | None]) -> list[object]:
    """Convert ``None`` values to blanks and coerce everything to strings/numbers."""

    sanitized: list[object] = []
    for value in row:
        if value is None:
            sanitized.append("")
        elif isinstance(value, (int, float)):
            sanitized.append(value)
        else:
            sanitized.append(str(value))
    return sanitized


def _apply_header_styles(worksheet: Worksheet) -> None:
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = CELL_BORDER


def _style_data_rows(worksheet: Worksheet) -> None:
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = DATA_ALIGNMENT
            cell.border = CELL_BORDER


def _auto_fit_columns(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(max_length + 6, 60)


def build_workbook_bytes(
    title: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[object | None]],
) -> bytes:
    """Return a byte-string representing an Excel workbook for the provided table."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = (title or "Export")[:31]
    worksheet.append([str(header) for header in headers])
    _apply_header_styles(worksheet)

    added_rows = False
    for row in rows:
        worksheet.append(_sanitize_row(row))
        added_rows = True

    if added_rows:
        _style_data_rows(worksheet)

    _auto_fit_columns(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


__all__ = ["build_workbook_bytes"]

